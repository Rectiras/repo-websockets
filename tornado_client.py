import os
import time
import json
import platform
import tornado.ioloop
import tornado.websocket
from tornado import gen
import openpyxl
import psutil

class WebSocketClient(tornado.websocket.WebSocketClientConnection):
    def on_message(self, message):
        print(f"Received from server: {message}")

        data = json.loads(message)

@gen.coroutine
def on_connect_future(future):
    try:
        ws = yield future
        print("Connected to WebSocket server")
        yield send_messages(ws)

    except Exception as e:
        print(f"Error connecting to WebSocket server: {e}")
        tornado.ioloop.IOLoop.current().stop()

@gen.coroutine
def send_messages(ws):
    latency_data = []
    latest_throughput = None  # Variable to store the latest throughput value

    while True:
        # send a message to the server
        message = input("Enter a message to send to the server (or 'exit' to quit): ")
        if message.lower() == 'exit':
            ws.close()
            break

        start_time = time.time()
        ws.write_message(message)
        real_time_data = yield ws.read_message()
        end_time = time.time()

        latency = end_time - start_time
        data = json.loads(real_time_data)

        print(f"Received from server:")
        print(f"  Data Type: {data['data_type']}")
        if "data_size" in data:
            data_size = data["data_size"]
            print(f"Data Size: {data_size} bytes")
        else:
            print("Data Size not available for the message.")
        print(f"Latency: {latency} seconds")

        if data["data_type"] == "throughput":
            throughput = data["data"]["throughput"]
            print(f"Throughput from server: {throughput} messages/second")
            # Update the latest throughput value
            latest_throughput = throughput

        # Append latency to the list
        latency_data.append(latency)

    # Write latency data, latest throughput, and hardware details to Excel file
    write_to_excel(latency_data, latest_throughput, data['data_type'], data_size)


def get_hardware_details():
    # Obtain hardware details using psutil
    cpu_info = f"CPU: {psutil.cpu_percent()}% usage"
    ram_info = f"RAM: {round(psutil.virtual_memory().total / (1024 ** 3), 2)} GB"
    os_info = f"OS: {platform.system()} {platform.version()}"
    return cpu_info, ram_info, os_info


def write_to_excel(latency_data, throughput=None, data_type=None, data_size=None):
    # Check if the file exists
    file_exists = os.path.isfile("tornado.xlsx")

    # Create a new workbook or load the existing one
    workbook = openpyxl.Workbook()

    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Latency"
    sheet["B1"] = "Data Type"
    sheet["C1"] = "Data Size in Bytes"
    sheet["D1"] = "Throughput (msg/s)"
    sheet["E1"] = "Average Latency"
    sheet["F1"] = "CPU"
    sheet["G1"] = "RAM"
    sheet["H1"] = "OS"

    # Get hardware details
    cpu_info, ram_info, os_info = get_hardware_details()

    # Write data to the sheet
    for row_num, latency in enumerate(latency_data, start=1):
        sheet.cell(row=row_num + 1, column=1, value=latency)

    # Write data type and data size
    sheet.cell(row=2, column=2, value=data_type)
    sheet.cell(row=2, column=3, value=data_size)

    # Write throughput only if it is not None
    if throughput is not None:
        sheet.cell(row=2, column=4, value=throughput)

    # Calculate average latency
    if latency_data:
        average_latency = sum(latency_data) / len(latency_data)
        sheet.cell(row=2, column=5, value=average_latency)

    # Write hardware details
    sheet.cell(row=2, column=6, value=cpu_info)
    sheet.cell(row=2, column=7, value=ram_info)
    sheet.cell(row=2, column=8, value=os_info)

    # Save the workbook
    workbook.save("tornado.xlsx")

if __name__ == "__main__":
    url = "ws://localhost:8888/ws"
    client = tornado.websocket.websocket_connect(url)
    client.add_done_callback(on_connect_future)
    tornado.ioloop.IOLoop.current().start()
