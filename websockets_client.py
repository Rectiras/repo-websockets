import asyncio
import json
import platform
import openpyxl
import os
import time
import psutil
import websockets

async def send_messages(uri):
    latency_data = []
    latest_throughput = None

    async with websockets.connect(uri) as websocket:
        while True:
            message = input("Enter a message to send to the server (or 'exit' to quit): ")
            if message.lower() == 'exit':
                break

            start_time = time.time()
            await websocket.send(message)
            real_time_data = await websocket.recv()
            end_time = time.time()

            latency = end_time - start_time
            data = json.loads(real_time_data)

            print(f"Received from server:")
            print(f"  Data Type: {data['data_type']}")
            if "data_size" in data:
                data_size = data["data_size"]
                print(f"Data Size: {data_size} bytes")
            else:
                print("Data Size not available for the message.")
            print(f"Latency: {latency} seconds")

            if data["data_type"] == "throughput":
                throughput = data["data"]["throughput"]
                print(f"Throughput from server: {throughput} messages/second")
                latest_throughput = throughput

            latency_data.append(latency)

    # Write latency data, latest throughput, and hardware details to Excel file
    await write_to_excel(latency_data, latest_throughput, data['data_type'], data_size)


async def write_to_excel(latency_data, throughput=None, data_type=None, data_size=None):
    file_exists = os.path.isfile("websockets.xlsx")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Latency"
    sheet["B1"] = "Data Type"
    sheet["C1"] = "Data Size in Bytes"
    sheet["D1"] = "Throughput (msg/s)"
    sheet["E1"] = "Average Latency"
    sheet["F1"] = "CPU"
    sheet["G1"] = "RAM"
    sheet["H1"] = "OS"

    cpu_info, ram_info, os_info = get_hardware_details()

    for row_num, latency in enumerate(latency_data, start=1):
        sheet.cell(row=row_num + 1, column=1, value=latency)

    sheet.cell(row=2, column=2, value=data_type)
    sheet.cell(row=2, column=3, value=data_size)

    if throughput is not None:
        sheet.cell(row=2, column=4, value=throughput)

    if latency_data:
        average_latency = sum(latency_data) / len(latency_data)
        sheet.cell(row=2, column=5, value=average_latency)

    sheet.cell(row=2, column=6, value=cpu_info)
    sheet.cell(row=2, column=7, value=ram_info)
    sheet.cell(row=2, column=8, value=os_info)

    workbook.save("websockets.xlsx")


def get_hardware_details():
    cpu_info = f"CPU: {psutil.cpu_percent()}% usage"
    ram_info = f"RAM: {round(psutil.virtual_memory().total / (1024 ** 3), 2)} GB"
    os_info = f"OS: {platform.system()} {platform.version()}"
    return cpu_info, ram_info, os_info


if __name__ == "__main__":
    asyncio.get_event_loop().run_until_complete(send_messages("ws://localhost:8888/ws"))
